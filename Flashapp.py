from flask import Flask, request, render_template
import pandas as pd
from openpyxl import load_workbook

app = Flask(__name__)

EXCEL_FILE = "mealData.xlsx"

@app.route('/')
def index():
    return render_template('form.html')

@app.route('/submit', methods=['POST'])
def submit():
    # Extracting data from the form
    data = {
        'Name': request.form['name'],
        'Days': request.form['gymDays'],
        'Time in gym': request.form['gymtime'],
        'Goal': request.form['fitnessTarget'],
        'Diet?': request.form.get('diet'),
        'What Diet': request.form['dietcurrent'],
        'Diet Strictness': request.form['dietStrictness']
    }

# this will convert is to dataframe ha    
    df = pd.DataFrame([data])

    try:
        workbook = load_workbook(EXCEL_FILE)
        sheet = workbook.active

        next_row = sheet.max_row + 1

        for col_num, value in enumerate(data.values(), start=1):
            sheet.cell(row=next_row, column=col_num, value=value)

        workbook.save(EXCEL_FILE)
        workbook.close()
    except FileNotFoundError:
        # will create a new excel file if it doesnt exist. Dunno if it going to work 
        df.to_excel(EXCEL_FILE, index=False)

    return '''
        <h2>Form Submitted Successfully!</h2>
        <p>Your response has been recorded.</p>
        <a href="/" style="display: inline-block; padding: 10px 15px; background: #007BFF; color: white; text-decoration: none; border-radius: 4px;">Submit Another Form</a>
    '''

if __name__ == '__main__':
    app.run(debug=True)
